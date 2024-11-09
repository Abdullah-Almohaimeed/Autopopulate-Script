from openpyxl import load_workbook  # Reading and writing without losing formatting
from rapidfuzz import process  # For approximate name matching
import pandas as pd  # Data manipulation
import os  # For checking file existence

# Automatically find one .xlsx and one .csv file in the current directory, assumed to be the main file (xlsx) and the data file (csv)
def find_files_in_directory():
    xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]

    if len(xlsx_files) != 1 or len(csv_files) != 1:
        print("Error: Ensure there is exactly one .xlsx and one .csv file in the directory.")
        exit(1)

    return xlsx_files[0], csv_files[0]

# Get the file names
MAIN_FILE, DATA_FILE = find_files_in_directory()

# Get the column names from the user
COLUMN_TO_UPDATE = input("Enter the name of the column that will be updated: ").strip()
COLUMN_TO_RIP_FROM = input("Enter the name of the column that data will be ripped from: ").strip()

print("Processing...")  # Very important for the user to know that the script is running

try:
    preserve_format = load_workbook(MAIN_FILE)
except Exception as e:
    print(f"Error: {e}") # Still don't know what could go wrong
    exit(1)

# PLEASE MAKE SURE TO INCLUDE THIS LINE. IF THE SHEET IS HIDDEN, IT WILL CORRUPT THE FILE
main_sheet = preserve_format.active  # Adjust the sheet name according to your data
if main_sheet.sheet_state != 'visible':
    main_sheet.sheet_state = 'visible'

# Get the header names and their column indexes
header = {cell.value: idx for idx, cell in enumerate(next(main_sheet.iter_rows(min_row=1, max_row=1)), start=0)}

# Use the first column as the matching column, first column header will be assumed to be the matching column
file_matcher_main = main_sheet.cell(1, 1).value  # First column header

# Load main dataset
dataframe_main = pd.read_excel(MAIN_FILE)

# Load the dataset that contains the data for the main file
dataframe_data = pd.read_csv(DATA_FILE)

# First column header will be assumed to be the matching column
file_matcher_data = dataframe_data.columns[0]  # First column header of the data file

# Find the best approximate match. Using rapidfuzz instead of fuzzywuzzy for performance
def find_best_match(row, lookup_df, lookup_col, return_col, threshold=80):
    match, *_ = process.extractOne(row.lower(), lookup_df[lookup_col].str.lower(), score_cutoff=threshold)
    if match:
        return lookup_df.loc[lookup_df[lookup_col].str.lower() == match, return_col].values[0]
    else:
        return None

# Apply the function to populate missing data based on approximate matches
dataframe_main[COLUMN_TO_UPDATE] = dataframe_main.apply(
    lambda row: find_best_match(row[file_matcher_main], dataframe_data, file_matcher_data, COLUMN_TO_RIP_FROM)
    if pd.isna(row[COLUMN_TO_UPDATE])
    else row[COLUMN_TO_UPDATE],
    axis=1
)

# Save updated data frame to the main Excel sheet
matcher_column_index = header[file_matcher_main]
update_column_index = header[COLUMN_TO_UPDATE]

for row in main_sheet.iter_rows(min_row=2, max_row=main_sheet.max_row):
    main_value = row[matcher_column_index].value

    # Lookup the value in the main DataFrame based on the matching key
    try:
        matched_value = dataframe_main.loc[dataframe_main[file_matcher_main] == main_value, COLUMN_TO_UPDATE].values[0]
    except IndexError:
        matched_value = None

    if matched_value is not None:
        # Update the appropriate cell in the main sheet using the dynamic column index
        row[update_column_index].value = matched_value

# Save the workbook after updating
preserve_format.save(MAIN_FILE)
print('Done')
